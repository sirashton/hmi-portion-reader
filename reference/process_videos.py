"""Batch entry point for agents: HMI slicer video(s) in -> portion-weights CSV(s) out.

For each matching video this runs the validated staged reader (read_video.py)
and writes `<video>_portion_weights.csv` next to it, in the analysis schema:
  Video title, Track, Log number, Portion number, Target Weight (g), Weight (g), Heel/End, Notes

Usage:
  # one video
  python process_videos.py "path/to/T2_Trial1.mp4" --target 667

  # every video in a folder starting with a prefix
  python process_videos.py --dir "path/to/folder" --prefix T2_ --target 667

  # target per video from TrialLog.xlsx (Trial Log sheet, video->target), if present
  python process_videos.py --dir "path/to/folder" --prefix T2_

Behaviour:
  - skips videos whose output CSV already exists (use --force to redo)
  - prints one PROGRESS/RESULT line per video and a final SUMMARY line
  - exit code 0 = all attempted videos produced a CSV; 1 = at least one failed

Requirements: pip install rapidocr-onnxruntime opencv-python numpy
              (optional, for TrialLog target lookup: openpyxl)
"""
import argparse, glob, os, re, sys, time

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import read_video  # noqa: E402


def triallog_targets(folder):
    """video-name -> target weight from TrialLog.xlsx (searched in the video
    folder, then up to two parent folders)."""
    path = None
    probe = os.path.abspath(folder)
    for _ in range(3):
        cand = os.path.join(probe, 'TrialLog.xlsx')
        if os.path.exists(cand):
            path = cand
            break
        probe = os.path.dirname(probe)
    if not path:
        return {}
    try:
        import openpyxl
    except ImportError:
        print('NOTE: TrialLog.xlsx present but openpyxl not installed; '
              'pass --target instead or pip install openpyxl', flush=True)
        return {}
    out = {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Trial Log' not in wb.sheetnames:
        return {}
    norm = lambda s: re.sub(r'[^a-z0-9]', '', str(s).lower())
    for r in wb['Trial Log'].iter_rows(min_row=5, values_only=True):
        if r[1] and r[8] and isinstance(r[5], (int, float)):
            out[norm(r[8])] = float(r[5])
    return out


def process_one(video, target, step, force):
    title = os.path.splitext(os.path.basename(video))[0]
    out = os.path.join(os.path.dirname(video), title + '_portion_weights.csv')
    if os.path.exists(out) and not force:
        print(f'SKIP   {title}: {os.path.basename(out)} already exists (--force to redo)', flush=True)
        return 'skip'
    print(f'START  {title} (target={target}g, step={step}s)', flush=True)
    t0 = time.time()
    try:
        store = read_video.run(video, 0, None, step, target)
        read_video.emit(store, title, target, out)
    except Exception as e:
        print(f'FAIL   {title}: {type(e).__name__}: {e}', flush=True)
        return 'fail'
    n = sum(1 for _ in open(out, encoding='utf-8')) - 1
    print(f'RESULT {title}: {n} portion rows -> {out}  ({time.time()-t0:.0f}s)', flush=True)
    return 'ok' if n > 0 else 'fail'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('video', nargs='?', help='single video file (or use --dir/--prefix)')
    ap.add_argument('--dir', default=None, help='folder to scan for videos')
    ap.add_argument('--prefix', default='', help='only videos whose filename starts with this')
    ap.add_argument('--target', type=float, default=None,
                    help='target portion weight in g (fallback: TrialLog.xlsx lookup)')
    ap.add_argument('--step', type=float, default=2.0, help='sampling interval, seconds')
    ap.add_argument('--force', action='store_true', help='reprocess even if the CSV exists')
    a = ap.parse_args()

    if a.video:
        vids = [a.video]
    elif a.dir:
        vids = sorted(v for v in glob.glob(os.path.join(a.dir, '*.mp4'))
                      if os.path.basename(v).startswith(a.prefix))
    else:
        ap.error('give a video file, or --dir (optionally with --prefix)')
    if not vids:
        print(f'SUMMARY no videos matched prefix "{a.prefix}" in {a.dir}')
        sys.exit(1)

    folder = a.dir or os.path.dirname(os.path.abspath(vids[0]))
    tl = triallog_targets(folder)
    norm = lambda s: re.sub(r'[^a-z0-9]', '', s.lower())

    counts = {'ok': 0, 'skip': 0, 'fail': 0}
    for v in vids:
        title = os.path.splitext(os.path.basename(v))[0]
        out = os.path.join(os.path.dirname(v), title + '_portion_weights.csv')
        if os.path.exists(out) and not a.force:      # skip before demanding a target
            print(f'SKIP   {title}: {os.path.basename(out)} already exists (--force to redo)', flush=True)
            counts['skip'] += 1
            continue
        target = a.target if a.target is not None else tl.get(norm(title))
        if target is None:
            print(f'FAIL   {title}: no target weight (pass --target or add the video '
                  f'to TrialLog.xlsx)', flush=True)
            counts['fail'] += 1
            continue
        counts[process_one(v, target, a.step, a.force)] += 1

    print(f"SUMMARY ok={counts['ok']} skipped={counts['skip']} failed={counts['fail']} "
          f"of {len(vids)} matched", flush=True)
    sys.exit(0 if counts['fail'] == 0 else 1)


if __name__ == '__main__':
    main()
